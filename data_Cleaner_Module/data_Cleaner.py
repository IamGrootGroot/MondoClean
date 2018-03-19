# -*- coding: UTF-8 -*-
import openpyxl
import os
import sys
import csv
import secrets
import string
from collections import Counter
from datetime import datetime
from random import randint

class Cleaner:

    def __init__(self):
        """ Cleaner object initialization """
        self.sheetN = 0
        self.banned = ['.',' ','#N/A','#DIV/0','inconnu','?','NA','None']
        self.wbList = []
        self.pathList = []
        self.given = []
        self.taskBytes = 0
        self.maxBytes = 100
        self.lastFile = ''
        self.lastIndex = 0

    def openWB(self, key, path):
        """ Load and open cleaner workbook:
        1: Open main workbook @self.path
        2: Open workbook to be aggregated to main workbook
        3: Open open data workbook to be jointed to main workbook
        """
        # mods = {'basic': 1, 'aggreg': 2, 'joint': 3}
        try:
            if path.endswith('.xlsx') or path.endswith('.XLSX'):
                if key == 1:
                    if path not in self.pathList:
                        self.wb = openpyxl.load_workbook(path)
                        self.wbC = openpyxl.load_workbook(path)
                        self.wbList.append(self.wbC)
                        self.pathList.append(path)
                    elif (path in self.pathList) and (len(self.wbList) == len(self.pathList)):
                        pass
                    else:
                        self.lastIndex = self.wbList.index(self.wbC)
                        self.wbC = openpyxl.load_workbook(path)
                        self.wbList.append(self.wbC)
                if key == 2:
                    self.wb1 = openpyxl.load_workbook(path)
                if key == 3:
                    self.wb2 = openpyxl.load_workbook(path)
                return ''
            else:
                raise ValueError("Le format de fichier n'est pas supporté. L'application supporte uniquement les formats: .xlsx,.xlsm,.xltx,.xltm")
        except ValueError as valerr:
            self.taskBytes = 0
            return valerr
        except FileNotFoundError as filerr:
            self.taskBytes = 0
            return filerr
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def anonymize(self, colIndexAN):
        """ Anonymization function: Sets cell values of given columns to random integer value
        if a duplicate is found, the set value is the same as the one encountered before. A second file is
        created for tracability purpose. This xlsx file contains the anonimyzation table with given values
        in column 1 and original values in column 2.

        1(colIndexAN is 1 integer): Anonymize column @colIndexAN
        2(colIndex is list of integers): Anonymize columns @colIndexAN, the values @column2 in the generated
        file become a concatenation of the values for each cell.
        """
        try:
            self.wba = openpyxl.Workbook()
            sheet = self.wb.worksheets[self.sheetN]
            self.maxBytes = sheet.max_row
            self.taskBytes = 0
            seen = {}
            head = ''
            self.wba.active.cell(row=1, column=1).value = 'ID'
            for s in colIndexAN:
                if s>sheet.max_column:
                    raise IndexError("L'index spécifié est supérieur aux nombres de colonnes du fichier")
                if type(s) is not int:
                    raise TypeError("Le type pour l'index n'est pas respecté")
                else:
                    head += '+'+str(sheet.cell(row=1, column=s).value)
            self.wba.active.cell(row=1, column=2).value = head
            if len(colIndexAN) == 1:
                for i in range(2, sheet.max_row+1):
                    if sheet.cell(row=i, column=colIndexAN[0]).value not in seen.values():
                        giv = randint(0, sheet.max_row*sheet.max_column)
                        while giv in self.given:
                            giv = randint(0, sheet.max_row*sheet.max_column)
                        seen.update({giv:sheet.cell(row=i, column=colIndexAN[0]).value})
                        self.wba.active.cell(row=self.wba.active.max_row+1, column=1).value = giv
                        self.wba.active.cell(row=self.wba.active.max_row, column=2).value = sheet.cell(row=i, column=colIndexAN[0]).value
                        self.wba.active.cell(row=self.wba.active.max_row, column=2).number_format = '0.00'
                        sheet.cell(row=i, column=colIndexAN[0]).value = giv
                        sheet.cell(row=i, column=colIndexAN[0]).number_format = '0.00'
                        self.given.append(giv)
                    else:
                        sheet.cell(row=i, column=colIndexAN[0]).value = list(seen.keys())[list(seen.values()).index(sheet.cell(row=i, column=colIndexAN[0]).value)]
                        sheet.cell(row=i, column=colIndexAN[0]).number_format = '0.00'
                    self.taskBytes = i-1
                colHead = {cell.value for n, cell in enumerate(list(sheet.rows)[0]) if n+1 == colIndexAN[0]}
            else:
                for i in range(2, sheet.max_row+1):
                    sequence = ''
                    for k in colIndexAN:
                        sequence += ' '+str(sheet.cell(row=i, column=k).value)
                    if sequence not in seen.values():
                        giv = randint(0, sheet.max_row*sheet.max_column)
                        while giv in self.given:
                            giv = randint(0, sheet.max_row*sheet.max_column)
                        for u in colIndexAN:
                            sheet.cell(row=i, column=u).value = giv
                            sheet.cell(row=i, column=u).number_format = '0.00'
                        seen.update({giv:sequence})
                        self.wba.active.cell(row=self.wba.active.max_row+1, column=1).value = giv
                        self.wba.active.cell(row=self.wba.active.max_row, column=2).value = sequence
                        self.wba.active.cell(row=self.wba.active.max_row, column=2).number_format = '0.00'
                        self.given.append(giv)
                    else:
                        for u in colIndexAN:
                            sheet.cell(row=i, column=u).value = list(seen.keys())[list(seen.values()).index(sequence)]
                            sheet.cell(row=i, column=u).number_format = '0.00'
                colIndexminus = []
                for h in colIndexAN:
                        colIndexminus.append(h-1)
                colHeads = {cell.value for n, cell in enumerate(list(sheet.rows)[0]) if n in colIndexminus}
                self.taskBytes = i-1
            return ''
            self.taskBytes = 0
        except ValueError:
            self.taskBytes = 0
            return "Erreur de la lecture des index: "+str(colIndexAN)+", vérifier que l'index spécifié est un nombre."
        except TypeError as typerr:
            self.taskBytes = 0
            return typerr
        except IndexError as indexerr:
            self.taskBytes = 0
            return indexerr
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def purify(self):
        """Purification function, removes banned characters given by the self.banned list"""
        try :
            sheet = self.wb.worksheets[self.sheetN]
            self.maxBytes = sheet.max_column*sheet.max_row
            self.taskBytes = 0
            track = 0
            for i in range(1, sheet.max_column+1):
                for j in range(2, sheet.max_row+1):
                    if type(sheet.cell(row=j, column=i).value) is str:
                        sheet.cell(row=j, column=i).value = sheet.cell(row=j, column=i).value.strip()
                    if str(sheet.cell(row=j, column=i).value) in self.banned:
                        sheet.cell(row=j, column=i).value = None
                        track=track+1
                    self.taskBytes=self.taskBytes+1
            self.taskBytes = 0
            return ''
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1]) + " at cell: [" + str(j) + " " + str(i) + "]"

    def formatNumbers(self):
        """Manages numbers formatting"""
        try:
            sheet = self.wb.worksheets[self.sheetN]
            self.maxBytes = sheet.max_column*sheet.max_row
            self.taskBytes = 0
            track = 0
            for i in range(1, sheet.max_column+1):
                for j in range(2, sheet.max_row+1):
                    if type(self.convertFloat(str(sheet.cell(row=j, column=i).value))) is float:
                        track=track+1
                        sheet.cell(row=j, column=i).value = float(sheet.cell(row=j, column=i).value)
                        sheet.cell(row=j, column=i).number_format = '0.00'
                    elif (type(self.convertFloat(str(sheet.cell(row=j, column=i).value)) is not float)) and ("," in str(sheet.cell(row=j, column=i).value)) and self.checkIsNumber(str(sheet.cell(row=j, column=i).value))=='number':
                        track=track+1
                        sheet.cell(row=j, column=i).value = float(str(sheet.cell(row=j, column=i).value).replace(',','.'))
                        sheet.cell(row=j, column=i).number_format = '0.00'
                    else:
                        pass
                    self.taskBytes=self.taskBytes+1
            self.taskBytes = 0
            return ''
        except :
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1]) + " at cell: [" + str(j) + " " + str(i) + "]"



    def changeDate(self, formatIn): ## TODO INDEX DATE
        """Reformats dates to the 'd/m/Y' format. The input format has to be specified by formatIn"""
        try :
            sheet = self.wb.worksheets[self.sheetN]
            self.maxBytes = len(list(sheet.rows)[0])*sheet.max_row
            self.taskBytes = 0
            dateStyleTag = self.dateHexGen()
            dateStyle = openpyxl.styles.NamedStyle(name=dateStyleTag, number_format='DD/MM/YYYY')
            self.wb.add_named_style(dateStyle)
            for n, head in enumerate(list(sheet.rows)[0]):
                dateKey = ["date", "Date", "DATE", "DT "]
                if any(key in head.value for key in dateKey):
                    for k in range(2, sheet.max_row+1):
                        if type(sheet.cell(row=k, column=n+1).value) is not datetime:
                            if str(sheet.cell(row=k, column=n+1).value) != '' and 'None' not in str(sheet.cell(row=k, column=n+1).value):
                                if ' ' in str(sheet.cell(row=k, column=n+1).value):
                                    sheet.cell(row=k, column=n+1).value = str(sheet.cell(row=k, column=n+1).value).replace(' ','')
                                if '.' in str(sheet.cell(row=k, column=n+1).value):
                                    sheet.cell(row=k, column=n+1).value = str(sheet.cell(row=k, column=n+1).value).replace('.','')
                                if '/' in str(sheet.cell(row=k, column=n+1).value):
                                    sheet.cell(row=k, column=n+1).value = str(sheet.cell(row=k, column=n+1).value).replace('/','')
                                if '-' in str(sheet.cell(row=k, column=n+1).value):
                                    sheet.cell(row=k, column=n+1).value = str(sheet.cell(row=k, column=n+1).value).replace('-','')
                                dateObject = datetime.strptime(str(sheet.cell(row=k, column=n+1).value),formatIn)
                                sheet.cell(row=k, column=n+1).value = dateObject.strptime(dateObject.strftime('%d/%m/%Y'),'%d/%m/%Y')
                                sheet.cell(row=k, column=n+1).style = dateStyleTag
                        else:
                            sheet.cell(row=k, column=n+1).value = sheet.cell(row=k, column=n+1).value.strptime(sheet.cell(row=k, column=n+1).value.strftime('%d/%m/%Y'),'%d/%m/%Y')
                            sheet.cell(row=k, column=n+1).style = dateStyleTag
                        self.taskBytes = self.taskBytes+1
                else:
                    self.taskBytes = self.taskBytes+sheet.max_row
            self.taskBytes = 0
            return ''
        except ValueError:
            self.taskBytes = 0
            return 'Error at cell['+str(k)+' '+str(n+1)+'], invalid cell content: ' + str(sheet.cell(row=k, column=n+1).value) + """Vérifier que le format d'entrée de la date correspond bien au format de date du fichier excel ."""
        except :
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def aggreg(self, paths):
        """Aggregation function, appends different xlsx files @paths, no need to worry about empty or missing columns.
        WARNING: Column names have to be the same as the ones in the main workbook"""
        try:
            for s in paths:
                if s.endswith('.xlsx') or s.endswith('.XLSX'):
                    continue
                else:
                    raise ValueError("Le format de fichier n'est pas supporté: "+s+".L'application supporte uniquement les formats: .xlsx,.xlsm,.xltx,.xltm")
            sheet = self.wb.worksheets[self.sheetN]
            self.maxBytes = len(paths)
            self.taskBytes = 0
            headers = list(sheet.rows)[0]
            for i, p in enumerate(paths):
                wbb = openpyxl.load_workbook(p)
                sheetB = wbb.active
                headersB = list(sheetB.rows)[0]
                patch_row = sheet.max_row
                for n, head in enumerate(headers):
                    for k, headB in enumerate(headersB):
                        if headB.value == head.value:
                            for j in range(2, sheetB.max_row+1):
                                sheet.cell(row=patch_row+j-1, column=n+1).value = sheetB.cell(row=j, column=k+1).value
                self.taskBytes = i+1
            self.taskBytes = 0
            return ''
        except ValueError as valerr:
            self.taskBytes = 0
            return valerr
        except FileNotFoundError as filerr:
            self.taskBytes = 0
            return filerr
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def param(self, listBans):
        """Add new banned characters"""
        try:
            self.banned = self.banned + listBans
            return ''
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

#Non Utilisé
    def lock(self):
        """Lock sheet"""
        try:
            if self.wb.protection.sheet == True:
                self.wb.protection.sheet = False
            else:
                self.wb.protection.sheet = True
            return ''
        except:
            self.taskBytes = 0
            return 'Locking failed.'

    def doublons(self, columns):
        """Identifies duplicates and highlights them"""
        try:
            uniq = {}
            dupes = []
            sheet = self.wb.worksheets[self.sheetN]
            for s in columns:
                if s>sheet.max_column:
                    raise IndexError("L'index spécifié est supérieur aux nombres de colonnes du fichier.")
                if type(s) is not int:
                    raise TypeError("Le type pour l'index n'est pas respecté.")
            self.maxBytes = len(columns)*sheet.max_row
            self.taskBytes = 0
            maxCol = sheet.max_column+1
            sheet.cell(row=1, column = maxCol).value = 'DOUBLON_FLAG'
            for row in sheet.iter_rows():
                sequence = ''
                for cell in row:
                    if cell.col_idx in columns:
                        sequence += str(cell.value)
                if sequence not in uniq.values():
                    uniq.update({cell.row:sequence})
                else:
                    dupes.append(cell.row)
                self.taskBytes = self.taskBytes+1
            for i in dupes:
                for j in range(1, sheet.max_column+1):
                    sheet.cell(row=i, column=j).fill = openpyxl.styles.PatternFill('solid', openpyxl.styles.colors.RED)
                    sheet.cell(row=i, column=maxCol).value = 'Doublon'
            self.taskBytes = 0
            return ''
        except ValueError:
            self.taskBytes = 0
            return "Erreur de la lecture des index: "+str(columns)+", vérifier que l'index spécifié est un nombre."
        except IndexError as indexerr:
            self.taskBytes = 0
            return indexerr
        except TypeError as typerr:
            self.taskBytes = 0
            return typerr
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def count(self, colIndex):
        """Counts the number of occurences of a given
        combination of column values and write that amount in the last column"""
        try:
            print('1')
            sheet = self.wb.worksheets[self.sheetN]
            for s in colIndex:
                if s>sheet.max_column:
                    raise IndexError("L'index spécifié est supérieur aux nombres de colonnes du fichier.")
                if type(s) is not int:
                    raise TypeError("Le type pour l'index n'est pas respecté.")
            self.maxBytes = sheet.max_row*2
            self.taskBytes = 0
            sequences = []
            colCount = sheet.max_column+1
            sheet.cell(row=1, column=colCount).value = 'FREQUENCE'
            for k, row in enumerate(sheet.iter_rows()):
                sequence = ''
                if k>0:
                    for n, cell in enumerate(row):
                        if n+1 in colIndex:
                            sequence += str(cell.value)
                    sequences.append(sequence)
                    self.taskBytes = self.taskBytes+1
            occurences = Counter(sequences)
            for n, s in enumerate(sequences):
                if s in occurences.keys():
                    sheet.cell(row=n+2, column=colCount).value = occurences.get(s)
                    self.taskBytes = self.taskBytes+1
            self.taskBytes = 0
            return ''
        except ValueError:
            self.taskBytes = 0
            return "Erreur de la lecture des index: "+str(colIndex)+", vérifier que l'index spécifié est un nombre."
        except IndexError as indexerr:
            self.taskBytes = 0
            return indexerr
        except TypeError as typerr:
            self.taskBytes = 0
            return typerr
        except :
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])


    def summ(self, colIndex, colAdd):
        """Sum of values @colAdd for same occurences of a given combination"""
        try:
            sheet = self.wb.worksheets[self.sheetN]
            for s in colIndex:
                if s>sheet.max_column:
                    raise IndexError("L'index spécifié pour le numéro de colonne à traiter est supérieur aux nombres de colonnes du fichier.")
                if type(s) is not int:
                    raise TypeError("Le type pour l'index n'est pas respecté.")
            if colAdd>sheet.max_column:
                    raise IndexError("L'index spécifié pour le numéro colonne à sommer est supérieur aux nombres de colonnes du fichier.")
            self.maxBytes = sheet.max_row*2
            self.taskBytes = 0
            sequences = {}
            colCount = sheet.max_column+1
            sheet.cell(row=1, column=colCount).value = 'SOMME'
            for k, row in enumerate(sheet.iter_rows()):
                sequence = ''
                if k>0:
                    for n, cell in enumerate(row):
                        if n+1 in colIndex:
                            sequence += str(cell.value)
                        if n+1==colAdd:
                            val = float(cell.value)   #Has to be numerical
                    if sequence in sequences:
                        sequences[sequence]+=val
                    else:
                        sequences.update({sequence:val})
                self.taskBytes = self.taskBytes+1
            for k, row in enumerate(sheet.iter_rows()):
                sequence = ''
                for n, cell in enumerate(row):
                    if n+1 in colIndex:
                        sequence += str(cell.value)
                if sequence in sequences:
                    sheet.cell(row=k+1, column=colCount).value = sequences.get(sequence)
                self.taskBytes = self.taskBytes+1
            self.taskBytes = 0
            return ''
        except ValueError:
            self.taskBytes = 0
            return "Erreur de la lecture des index, vérifier que l'index spécifié est un nombre."
        except IndexError as indexerr:
            self.taskBytes = 0
            return indexerr
        except TypeError as typerr:
            self.taskBytes = 0
            return typerr
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def joint(self, path, colComp1, colComp2, colJoints):
        """Joint opendata @path. Finds matching values between colComp1 and colComp2
        and add the data in colJoint at matching index"""
        try:
            if path.endswith('.xlsx') or path.endswith('.XLSX'):
                cleaner2 = Cleaner()
                cleaner2.openWB(1,path)
                cleaner2.purify()
                cleaner2.changeDate(None)
                sheetB = cleaner2.wb.worksheets[cleaner2.sheetN]
                sheet = self.wb.worksheets[self.sheetN]
                if colComp1>sheet.max_column:
                    raise IndexError("L'index spécifié pour le numéro de colonne à matcher du fichier 1 est supérieur aux nombres de colonnes du fichier.")
                if colComp2>sheetB.max_column:
                    raise IndexError("L'index spécifié pour le numéro de colonne à matcher du fichier 2 est supérieur aux nombres de colonnes du fichier.")
                for s in colJoints:
                    if s>sheetB.max_column:
                        raise IndexError("L'index spécifié pour le numéro de colonne à joindre est supérieur aux nombres de colonnes du fichier.")
                self.taskBytes = 0
                mr = sheet.max_row
                mb = sheetB.max_row
                mc = sheet.max_column
                self.maxBytes = mr+mb
                heads = []
                for row in sheetB.iter_rows(min_row=1, max_row=1):
                    for n, cell in enumerate(row):
                        if n+1 in colJoints:
                            heads.append(str(cell.value))
                colc1 = []
                colc2 = []
                colj = []
                for col1 in sheet.iter_cols(min_row=2, min_col=colComp1, max_col=colComp1, max_row=mr):
                    for cell1 in col1:
                        colc1.append(cell1.value)
                    self.taskBytes = self.taskBytes+1
                for col2 in sheetB.iter_cols(min_row=2, min_col=colComp2, max_col=colComp2, max_row=mb):
                    for i, cell2 in enumerate(col2):
                        joints = []
                        colc2.append(cell2.value)
                        for k in range(len(colJoints)):
                            for rowJ in sheetB.iter_rows(min_row=i+2, min_col=colJoints[k], max_col=colJoints[k], max_row=i+2):
                                for cell3 in rowJ:
                                    joints.append(cell3.value)
                        colj.append(joints)
                    self.taskBytes = self.taskBytes+1
                idx = {}
                self.taskBytes = 0
                self.maxBytes = len(colc1)*len(colc2)
                for i in range(len(colc1)):
                    for j in range(len(colc2)):
                        if colc1[i]==colc2[j]:
                            idx.update({i:colj[j]})
                        self.taskBytes = self.taskBytes+1
                self.taskBytes = 0
                self.maxBytes = len(idx.keys())*len(colJoints)
                for i, j in enumerate(idx.keys()):
                    for k in range(len(colJoints)):
                        sheet.cell(row=j+2, column=k+1+mc).value = idx.get(j)[k]
                        self.taskBytes = self.taskBytes+1
                self.taskBytes = 0
                for n, s in enumerate(heads):
                    sheet.cell(row=1, column=n+1+mc).value = s
                return ''
            else:
                raise ValueError("Le format de fichier n'est pas supporté. L'application supporte uniquement les formats: .xlsx,.xlsm,.xltx,.xltm")
        except ValueError as valerr:
            self.taskBytes = 0
            return valerr
        except FileNotFoundError as filerr:
            self.taskBytes = 0
            return filerr
        except IndexError as indexerr:
            self.taskBytes = 0
            return indexerr
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def categorize(self, mod, colIndexC, changes):
        """Catégorisation, changes cell values @column
        colIndexC to the corresponding key in the changes dict"""
        try:
            if mod == "numerical":
                sheet = self.wb.worksheets[self.sheetN]
                if colIndexC>sheet.max_column:
                    raise IndexError("L'index spécifié est supérieur aux nombres de colonnes du fichier.")
                maxCol = sheet.max_column+1
                self.maxBytes = sheet.max_row
                self.taskBytes = 0
                sheet.cell(row=1, column=maxCol).value = "CATEG "+str(sheet.cell(row=1, column=colIndexC).value)
                for col in sheet.iter_cols(min_row=2, min_col=colIndexC, max_col=colIndexC, max_row=sheet.max_row):
                    for k, cell in enumerate(col):
                        mask = None
                        if (cell.value not in self.banned) and (cell.value is not None):
                            mask = [mask for n, mask in enumerate(list(changes.values())) if int(list(changes.keys())[n][0]) <= int(cell.value) <= int(list(changes.keys())[n][-1])]
                            if mask:
                                sheet.cell(row=k+2, column=maxCol).value = mask[0]
                            else:
                                pass
                        else:
                            pass
                    self.taskBytes = self.taskBytes+1
                self.taskBytes = 0
            if mod == "substitute":
                sheet = self.wb.worksheets[self.sheetN]
                if colIndexC>sheet.max_column:
                    raise IndexError("L'index spécifié est supérieur aux nombres de colonnes du fichier.")
                maxCol = sheet.max_column+1
                self.maxBytes = sheet.max_row
                self.taskBytes = 0
                sheet.cell(row=1, column=maxCol).value = "CATEG "+str(sheet.cell(row=1, column=colIndexC).value)
                for col in sheet.iter_cols(min_row=2, min_col=colIndexC, max_col=colIndexC, max_row=sheet.max_row):
                    for k, cell in enumerate(col):
                        mask = None
                        if cell.value != 'None':
                            mask = [mask for n, mask in enumerate(list(changes.values())) if list(changes.keys())[n]==cell.value]
                            if mask:
                                sheet.cell(row=k+2, column=maxCol).value = mask[0]
                            else:
                                pass
                        else:
                            pass
                    self.taskBytes = self.taskBytes+1
                self.taskBytes = 0
            return ''
        except IndexError as indexerr:
            self.taskBytes = 0
            return indexerr
        except ValueError:
            self.taskBytes = 0
            return "Erreur de la lecture des index, vérifier que l'index spécifié est un nombre."
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def addIndex(self):
        try:
            sheet = self.wb.worksheets[self.sheetN]
            self.maxBytes = sheet.max_column*sheet.max_row
            self.taskBytes = 0
            track = 0;
            for row in sheet.iter_rows(max_row=1):
                for n, cell in enumerate(row):
                    l = len(str(n+1))
                    if str(cell.value)[0:l] != str(n+1):
                        cell.value = str(n+1)+' '+cell.value
                    else:
                        pass
            return ''
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def convertFloat(self, value):
        try:
            return float(value)
        except:
            self.taskBytes = 0
            return(value)

    def checkIsNumber(self, s):
        try:
            for i in s:
                int(i)
            return 'number'
        except:
            self.taskBytes = 0
            return 'not number'

    def timeMachine(self, request, *args):
        """A time machine to allow undo and resets"""
        try:
            if request == 'pullBack':
                #del self.wbList[-1]
                #os.remove(self.pathList[-1])####LES LIGNES QUI SUPPRIMENT
                #del self.pathList[-1]
                self.wb = openpyxl.load_workbook(self.pathList[self.lastIndex])
                self.lastIndex = self.lastIndex-1
                if self.lastIndex == -1:
                    self.lastIndex = 0
                return self.pathList[self.lastIndex]
            if request == 'pullBack@':
                self.wb = openpyxl.load_workbook(args[0])
                self.lastIndex = self.pathList.index(args[0])-1
                return args[0]
            if request == 'fullReset':
                #del self.wbList[1:]
                #for p in self.pathList[1:]: ####Suppriment aussi
                    #os.remove(p)
                #del self.pathList[1:]
                self.lastIndex = 0
                self.wb = openpyxl.load_workbook(self.pathList[0])
                return self.pathList[0]
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def dateHexGen(self):
        """Hex key generator for dateStyle"""
        try:
            return secrets.token_hex(32)
        except:
            self.taskBytes = 0
            return secrets.token_hex(32)

    def getProgress(self):
        return (self.taskBytes*100)/self.maxBytes

    def resetBanned(self):
        try:
            self.banned = ['.',' ','#N/A','#DIV/0','inconnu','?','NA','None']
            return ''
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])

    def saveWB(self, key, newPath):
        """Save workbook:
        1: Save main workbook @newPath
        2: Save both main workbook and anonymization @newPath and newPath+'_ID'
        """
        try:
            if key == 1:
                self.wb.save(newPath)
                self.pathList.append(newPath)
                return ''
            else:
                self.wb.save(newPath)
                self.pathList.append(newPath)
                index = newPath.find('.xlsx')
                pathWBA = newPath[:index] + '_ID' + newPath[index:]
                self.wb.save(newPath)
                self.wba.save(pathWBA)
                return ''
        except:
            self.taskBytes = 0
            return "Unexpected error: " + str(sys.exc_info()[0]) + str(sys.exc_info()[1])
